import openpyxl

class Excel(object):
    def __init__(self,path,target_sheet_name='Sheet1'):
        self.path=path
        self.workbook=None
        self.sheet=None
        self.row_num=None
        self.column_num=None
        self.column_names=None
        self.init(target_sheet_name)

    def init(self,target_sheet_name):
        self.workbook = openpyxl.load_workbook(self.path)
        self.sheet= self.workbook.get_sheet_by_name(target_sheet_name)
        self.row_num=self.sheet.max_row
        self.column_num=self.sheet.max_column
        self.column_names=list(self.sheet.row_values(0))

    def get_row_num(self):
        return self.row_num

    def get_column_num(self):
        return self.column_num

    def get_column_names(self):
        return self.column_names

    def get_row_values(self,row_num):
        return list(self.sheet.row_values(row_num))

    def get_column_values(self,column_num):
        return list(self.sheet)



data=Excel(path='/home/hyl/data/data-lyl/食品列表+标签500类-2018-10-24.xlsx')
print(data.get_row_num())
